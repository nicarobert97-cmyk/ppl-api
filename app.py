from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
import json, io

app = Flask(__name__)
CORS(app)

def fill(hex): return PatternFill("solid", fgColor=hex)
def font(hex, bold=True, size=11): return Font(name="Calibri", size=size, bold=bold, color=hex)
def align(h="center"): return Alignment(horizontal=h, vertical="center")

thin   = Side(style="thin",   color="000000")
medium = Side(style="medium", color="000000")
thin_b   = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
medium_b = Border(left=medium, right=medium, top=medium, bottom=medium)

def sc(cell, bg, fg, bold=True, border=None, h="center", size=11):
    if bg: cell.fill = fill(bg)
    cell.font      = font(fg, bold, size)
    cell.border    = border or thin_b
    cell.alignment = align(h)

WEEK_BG="0B2F4D"; SET1_BG="1F4E78"; SET2_BG="305496"
WHITE="FFFFFF";   DARK="111111";    EX_BG="FFFFFF"
SET_BG=["D9F2D9","FFF2CC","FCE4D6","F8CBAD"]
INC_BG="E2EFDA"; INC_FG="1F7A1F"
HLD_BG="FFF2CC"; HLD_FG="7F6000"
DEC_BG="FCE4D6"; DEC_FG="CC0000"

WEEK_LABELS=[
    "Week 1 (23 Feb - 01 Mar)","Week 2 (02 Mar - 08 Mar)","Week 3 (09 Mar - 15 Mar)",
    "Week 4 (16 Mar - 22 Mar)","Week 5 (23 Mar - 29 Mar)","Week 6 (30 Mar - 05 Apr)",
    "Week 7 (06 Apr - 12 Apr)","Week 8 (13 Apr - 19 Apr)","Week 9 (20 Apr - 26 Apr)",
    "Week 10 (27 Apr - 03 May)","Week 11 (04 May - 10 May)","Week 12 (11 May - 17 May)",
]

PUSH_EXERCISES = [
    {"name":"Incline Dumbbell Press",         "sets":4,"settings_row":2,"init":[26,26,24,24],
     "targets":[{"min":6,"max":8},{"min":8,"max":8},{"min":8,"max":10},{"min":8,"max":10}]},
    {"name":"Dumbbell Flyes",                 "sets":4,"settings_row":3,"init":[16,16,14,14],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":10,"max":12},{"min":10,"max":12}]},
    {"name":"Face Pulls",                     "sets":4,"settings_row":4,"init":[20,20,20,17.5],
     "targets":[{"min":12,"max":12},{"min":12,"max":12},{"min":12,"max":15},{"min":15,"max":15}]},
    {"name":"Cable Lateral Raises",           "sets":4,"settings_row":5,"init":[5,5,5,5],
     "targets":[{"min":12,"max":15},{"min":12,"max":15},{"min":12,"max":15},{"min":12,"max":12}]},
    {"name":"Straight Bar Triceps Extension", "sets":4,"settings_row":6,"init":[20,20,20,17.5],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":10},{"min":10,"max":12}]},
    {"name":"Overhead Rope Triceps Extension","sets":4,"settings_row":7,"init":[22.5,22.5,22.5,20],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":8},{"min":10,"max":12}]},
]
PULL_EXERCISES = [
    {"name":"Wide Grip Lat Pulldown",     "sets":3,"settings_row":8, "init":[55,55,55],
     "targets":[{"min":8,"max":12},{"min":8,"max":12},{"min":8,"max":12}]},
    {"name":"Close Grip Lat Pulldown",    "sets":3,"settings_row":9, "init":[55,55,55],
     "targets":[{"min":8,"max":12},{"min":8,"max":12},{"min":8,"max":12}]},
    {"name":"Seated Cable Row",           "sets":3,"settings_row":10,"init":[45,45,45],
     "targets":[{"min":8,"max":12},{"min":8,"max":12},{"min":8,"max":12}]},
    {"name":"Wide Grip Seated Cable Row", "sets":3,"settings_row":11,"init":[40,40,40],
     "targets":[{"min":8,"max":12},{"min":8,"max":12},{"min":8,"max":12}]},
    {"name":"Incline Dumbbell Curl",      "sets":3,"settings_row":12,"init":[10,10,10],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":10}]},
    {"name":"Machine Bicep Curl",         "sets":3,"settings_row":13,"init":[22.5,22.5,22.5],
     "targets":[{"min":10,"max":15},{"min":10,"max":15},{"min":10,"max":15}]},
]
LEGS_EXERCISES = [
    {"name":"Squats",             "sets":3,"settings_row":14,"init":[70,70,70],
     "targets":[{"min":6,"max":10},{"min":6,"max":10},{"min":6,"max":10}]},
    {"name":"Romanian Deadlift",  "sets":3,"settings_row":15,"init":[80,80,80],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":10}]},
    {"name":"Incline Leg Press",  "sets":3,"settings_row":16,"init":[200,200,200],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":10}]},
    {"name":"Seated Calf Raises", "sets":3,"settings_row":17,"init":[120,120,120],
     "targets":[{"min":12,"max":15},{"min":12,"max":15},{"min":12,"max":15}]},
    {"name":"Leg Extension",      "sets":3,"settings_row":18,"init":[85,85,85],
     "targets":[{"min":8,"max":10},{"min":8,"max":10},{"min":8,"max":10}]},
    {"name":"Leg Curl",           "sets":3,"settings_row":19,"init":[45,45,45],
     "targets":[{"min":8,"max":12},{"min":8,"max":12},{"min":8,"max":12}]},
]

def get_actual(data, wkey, week_idx, ex_name, si):
    try: return data[wkey][str(week_idx)][ex_name][str(si)]
    except: pass
    try: return data[wkey][week_idx][ex_name][si]
    except: return None

def get_decision(data, wkey, week_idx, ex):
    actuals = [get_actual(data, wkey, week_idx, ex["name"], si) for si in range(ex["sets"])]
    if any(a is None or a == "" for a in actuals): return ""
    actuals = [int(a) for a in actuals]
    fails = sum(1 for i,a in enumerate(actuals) if a < ex["targets"][i]["min"])
    if fails >= 2: return "DECREASE"
    if all(a >= ex["targets"][i]["max"] for i,a in enumerate(actuals)): return "INCREASE"
    return "HOLD"

def generate_excel(state):
    settings = state.get("settings", {})
    data     = state.get("data", {})

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Settings sheet
    ws = wb.create_sheet("Settings")
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.row_dimensions[1].height = 24
    for ci, h in enumerate(["Exercise","Step + (kg)","Step - (kg)"],1):
        c = ws.cell(row=1, column=ci, value=h)
        sc(c, SET1_BG, WHITE, border=medium_b)
    all_ex = PUSH_EXERCISES + PULL_EXERCISES + LEGS_EXERCISES
    for i, ex in enumerate(all_ex, 2):
        s = settings.get(ex["name"], {"up":2.5,"down":2.5})
        bg = "D9E1F2" if i%2==0 else EX_BG
        ws.row_dimensions[i].height = 20.1
        c1 = ws.cell(row=i, column=1, value=ex["name"])
        sc(c1, bg, DARK, h="left")
        c2 = ws.cell(row=i, column=2, value=s["up"])
        sc(c2, bg, DARK)
        c3 = ws.cell(row=i, column=3, value=s["down"])
        sc(c3, bg, DARK)

    def build_sheet(sheet_name, wkey, exercises):
        ws = wb.create_sheet(sheet_name)
        set_count = exercises[0]["sets"]
        set_start = [2, 6, 10, 14]
        last_vis_col = set_start[set_count-1] + 3
        hidden_start = last_vis_col + 1
        tmin_cols = list(range(hidden_start, hidden_start + set_count))
        tmax_cols = list(range(hidden_start + set_count, hidden_start + 2*set_count))
        decision_col = hidden_start + 2*set_count

        ws.column_dimensions["A"].width = 40
        for si in range(set_count):
            sc_ = set_start[si]
            ws.column_dimensions[get_column_letter(sc_)].width = 9
            ws.column_dimensions[get_column_letter(sc_+1)].width = 14
            ws.column_dimensions[get_column_letter(sc_+2)].width = 12
            ws.column_dimensions[get_column_letter(sc_+3)].width = 7
        ws.column_dimensions[get_column_letter(decision_col)].width = 13
        for col in tmin_cols + tmax_cols:
            ws.column_dimensions[get_column_letter(col)].hidden = True

        week_ex_rows = []
        row = 1

        for week_idx in range(12):
            ex_rows_this_week = []

            # Week header
            ws.row_dimensions[row].height = 26.1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_vis_col)
            c = ws.cell(row=row, column=1, value=WEEK_LABELS[week_idx])
            sc(c, WEEK_BG, WHITE, border=medium_b, h="left", size=13)
            for col in range(2, last_vis_col+1):
                sc(ws.cell(row=row, column=col), WEEK_BG, WHITE, border=medium_b)
            row += 1

            # Set group header
            ws.row_dimensions[row].height = 21.95
            set_hdr_row = row
            sc(ws.cell(row=row, column=1, value="Exercise"), SET1_BG, WHITE, border=medium_b)
            for si in range(set_count):
                sc_ = set_start[si]
                ws.merge_cells(start_row=row, start_column=sc_, end_row=row, end_column=sc_+3)
                sc(ws.cell(row=row, column=sc_, value=f"Set {si+1}"), SET1_BG, WHITE, border=medium_b)
                for x in range(1,4):
                    sc(ws.cell(row=row, column=sc_+x), SET1_BG, WHITE, border=medium_b)
            row += 1

            # Sub-header
            ws.row_dimensions[row].height = 21.95
            ws.merge_cells(start_row=set_hdr_row, start_column=1, end_row=row, end_column=1)
            sc(ws.cell(row=set_hdr_row, column=1), SET1_BG, WHITE, border=medium_b)
            sc(ws.cell(row=row, column=1), SET2_BG, WHITE, border=medium_b)
            for si in range(set_count):
                sc_ = set_start[si]
                for x, lbl in enumerate(["Kg","Target Reps","Actual Reps","Δ"]):
                    sc(ws.cell(row=row, column=sc_+x, value=lbl), SET2_BG, WHITE, border=medium_b)
            row += 1

            # Exercise rows
            for ex_idx, ex in enumerate(exercises):
                ws.row_dimensions[row].height = 20.1
                ex_rows_this_week.append(row)
                sr = ex["settings_row"]

                sc(ws.cell(row=row, column=1, value=ex["name"]), EX_BG, DARK, h="left")

                for si in range(ex["sets"]):
                    sc_ = set_start[si]
                    bg  = SET_BG[si]
                    tgt = ex["targets"][si]
                    tgt_str = str(tgt["min"]) if tgt["min"]==tgt["max"] else f'{tgt["min"]}-{tgt["max"]}'

                    kg_cell = ws.cell(row=row, column=sc_)
                    if week_idx == 0:
                        kg_cell.value = ex["init"][si]
                    else:
                        pr = week_ex_rows[week_idx-1][ex_idx]
                        kg_let  = get_column_letter(sc_)
                        dec_let = get_column_letter(decision_col)
                        kg_cell.value = (
                            f'=IF({dec_let}{pr}="INCREASE",{kg_let}{pr}+Settings!$B${sr},'
                            f'IF({dec_let}{pr}="DECREASE",{kg_let}{pr}-Settings!$C${sr},{kg_let}{pr}))'
                        )
                    sc(kg_cell, bg, DARK)

                    sc(ws.cell(row=row, column=sc_+1, value=tgt_str), bg, DARK)

                    actual_val = get_actual(data, wkey, week_idx, ex["name"], si)
                    c = ws.cell(row=row, column=sc_+2, value=int(actual_val) if actual_val not in (None,"") else None)
                    sc(c, bg, DARK)

                    act_let  = get_column_letter(sc_+2)
                    tmax_let = get_column_letter(tmax_cols[si])
                    delta_cell = ws.cell(row=row, column=sc_+3)
                    delta_cell.value = f'=IF({act_let}{row}="","",{act_let}{row}-${tmax_let}{row})'
                    sc(delta_cell, bg, DARK)

                for si in range(ex["sets"]):
                    ws.cell(row=row, column=tmin_cols[si]).value = ex["targets"][si]["min"]
                for si in range(ex["sets"]):
                    ws.cell(row=row, column=tmax_cols[si]).value = ex["targets"][si]["max"]

                act_refs  = [get_column_letter(set_start[si]+2)+str(row) for si in range(ex["sets"])]
                tmin_refs = ["$"+get_column_letter(tmin_cols[si])+str(row) for si in range(ex["sets"])]
                tmax_refs = ["$"+get_column_letter(tmax_cols[si])+str(row) for si in range(ex["sets"])]
                fail_sum  = "+".join([f"IF({a}<{mn},1,0)" for a,mn in zip(act_refs,tmin_refs)])
                all_max   = ",".join([f"{a}>={mx}" for a,mx in zip(act_refs,tmax_refs)])
                dec_formula = (
                    f'=IF(COUNTA({",".join(act_refs)})<{ex["sets"]},"", '
                    f'IF(({fail_sum})>=2,"DECREASE",'
                    f'IF(AND({all_max}),"INCREASE","HOLD")))'
                )
                dec_cell = ws.cell(row=row, column=decision_col)
                dec_cell.value = dec_formula

                dec = get_decision(data, wkey, week_idx, ex)
                if dec == "INCREASE":   dec_bg, dec_fg = INC_BG, INC_FG
                elif dec == "HOLD":     dec_bg, dec_fg = HLD_BG, HLD_FG
                elif dec == "DECREASE": dec_bg, dec_fg = DEC_BG, DEC_FG
                else:                   dec_bg, dec_fg = EX_BG,  DARK
                sc(dec_cell, dec_bg, dec_fg)

                row += 1

            week_ex_rows.append(ex_rows_this_week)
            ws.row_dimensions[row].height = 8
            row += 1

    build_sheet("Push Plan", "push", PUSH_EXERCISES)
    build_sheet("Pull Plan", "pull", PULL_EXERCISES)
    build_sheet("Legs Plan", "legs", LEGS_EXERCISES)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

@app.route('/export', methods=['POST', 'OPTIONS'])
def export():
    if request.method == 'OPTIONS':
        return '', 204
    try:
        state = request.get_json()
        output = generate_excel(state)
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name='Push_Pull_Legs_Tracker.xlsx'
        )
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route('/')
def health():
    return 'PPL Tracker API is running!'

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000)
